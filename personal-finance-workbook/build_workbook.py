"""
Builds Joel's Personal Finance Workbook (.xlsx) — a single dark-themed Excel
workbook with native charts, dropdowns, and input-driven formulas.

Run with:  python3 build_workbook.py
Produces:  Personal_Finance_Workbook.xlsx

Design rules baked into every sheet:
  * One dark navy theme, one consistent "input cell" colour (soft gold).
  * The user only ever types into gold cells — everything else is a formula.
  * All charts are native Excel chart objects bound to live cell ranges.
"""

import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl.chart import (
    BarChart, LineChart, AreaChart, DoughnutChart, PieChart, Reference, Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.title import Title
from openpyxl.chart.text import Text, RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties, RegularTextRun,
    RichTextProperties, Font as DFont,
)
from openpyxl.formatting.rule import FormulaRule, DataBarRule, CellIsRule

# ════════════════════════════════════════════════════════════════════════
#  PALETTE  —  one dark theme shared by every sheet
# ════════════════════════════════════════════════════════════════════════
NAVY      = "1A1E2B"     # sheet background (deep blue-charcoal)
CARD      = "232A3D"     # card / panel background
CARD2     = "2B3349"     # slightly lighter card (table headers, alt rows)
CARD3     = "313A54"     # hover / strong card / chart plot area
GRID      = "3C4660"     # gridlines & borders (subtle)
TEAL      = "35E6B6"     # mint/teal — income, savings, positive, "good"
TEAL_DK   = "1FB593"
PINK      = "FF6792"     # magenta/pink — expenses, debt, negative
PINK_DK   = "E8497C"
BLUE      = "5B93FF"     # accent — charts / links
YELLOW    = "FFCF5C"     # accent — headers / highlights
PURPLE    = "B084F5"     # accent — charts
WHITE     = "F4F6FB"     # primary text
GREY      = "AEB7CF"     # body / label text
GREY_DK   = "76819B"     # muted / helper text
ROSE      = "FF8FAE"     # soft pink for subtle warnings

GOLD_FILL   = "F6D998"   # ===> THE one-and-only INPUT CELL colour (soft gold)
GOLD_TEXT   = "1A1E2B"   # dark navy text sits on the gold fill
GOLD_BORDER = "E7B64C"   # gold border that frames every input cell

FMT_CAD   = '"$"#,##0.00'
FMT_CAD0  = '"$"#,##0'
FMT_PCT   = '0.0%'
FMT_PCT0  = '0%'
FMT_DATE  = 'mmm d, yyyy'
FMT_MONY  = 'mmm yyyy'
FMT_NUM2  = '#,##0.00'
FMT_NUM1  = '#,##0.0'

FONT_NAME = "Calibri"

# ════════════════════════════════════════════════════════════════════════
#  LOW-LEVEL STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════

def fl(color):
    """Solid PatternFill from a hex string."""
    return PatternFill(fill_type="solid", fgColor=color)


def fnt(color=GREY, size=10, bold=False, italic=False, name=FONT_NAME):
    return Font(name=name, color=color, size=size, bold=bold, italic=italic)


CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
LEFT_I = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT  = Alignment(horizontal="right", vertical="center")
WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_C  = Alignment(horizontal="center", vertical="top", wrap_text=True)


def side(color=GRID, style="thin"):
    return Side(style=style, color=color)


def box(color=GRID, style="thin"):
    s = side(color, style)
    return Border(left=s, right=s, top=s, bottom=s)


def border_bottom(color=GRID, style="thin"):
    return Border(bottom=side(color, style))


def border_left(color, style="medium"):
    return Border(left=side(color, style))


def setc(ws, addr, value=None, fill=None, color=GREY, size=10, bold=False,
         italic=False, align=None, fmt=None, border=None, name=FONT_NAME):
    """Set a cell's value + style in one call. `addr` is either 'B3' or (row, col). Returns the cell."""
    if isinstance(addr, tuple):
        c = ws.cell(row=addr[0], column=addr[1])
    else:
        c = ws[addr]
    if value is not None:
        c.value = value
    c.font = fnt(color, size, bold, italic, name)
    if fill is not None:
        c.fill = fl(fill)
    if align is not None:
        c.alignment = align
    if fmt is not None:
        c.number_format = fmt
    if border is not None:
        c.border = border
    return c


def paint_bg(ws, rows=80, cols=24, color=NAVY):
    """Paint the whole sheet's working area with the dark background colour."""
    f = fl(color)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = f


def rect(ws, r1, c1, r2, c2, color=CARD):
    """Fill a rectangular block of cells solid."""
    f = fl(color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f


def accent_bar(ws, row, c1, c2, color):
    f = fl(color)
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = f


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return ws.cell(row=r1, column=c1)


def title_block(ws, row, col, text, subtitle=None, color=WHITE, accent=TEAL):
    """Page title with a small accent chip to its left."""
    setc(ws, (row, col), value="", fill=accent)
    cell = merge(ws, row, col + 1, row, col + 9)
    setc(ws, cell.coordinate, value=text, color=color, size=20, bold=True)
    cell.alignment = LEFT_I
    if subtitle:
        sc = merge(ws, row + 1, col + 1, row + 1, col + 9)
        setc(ws, sc.coordinate, value=subtitle, color=GREY, size=10.5, italic=True)
        sc.alignment = LEFT_I


def section_header(ws, row, col, text, color=TEAL, span=8):
    cell = merge(ws, row, col, row, col + span - 1)
    setc(ws, cell.coordinate, value=text.upper(), color=color, size=12.5, bold=True)
    cell.alignment = LEFT_I


def note(ws, row, col, text, span=8, height=None):
    cell = merge(ws, row, col, row, col + span - 1)
    setc(ws, cell.coordinate, value="ⓘ  " + text, color=GREY_DK, size=9, italic=True, align=WRAP)
    if height:
        ws.row_dimensions[row].height = height
    return cell


def label(ws, addr, text, color=GREY, size=9.5, bold=False, align=LEFT_I):
    return setc(ws, addr, value=text, color=color, size=size, bold=bold, align=align)


def value_cell(ws, addr, value=None, fmt=FMT_CAD, color=WHITE, size=11, bold=True, align=LEFT_I, border=None):
    return setc(ws, addr, value=value, color=color, size=size, bold=bold, fmt=fmt, align=align, border=border)


def input_cell(ws, addr, value=None, fmt=None, align=CENTER):
    """THE input style — soft gold fill, dark bold text, gold border. Used everywhere."""
    c = setc(ws, addr, value=value, fill=GOLD_FILL, color=GOLD_TEXT, size=10.5,
             bold=True, align=align, fmt=fmt, border=box(GOLD_BORDER))
    return c


def tick_cell(ws, addr):
    """A tickable input cell (data-validation list is attached separately)."""
    c = setc(ws, addr, value="", fill=GOLD_FILL, color=TEAL_DK, size=11, bold=True,
             align=CENTER, border=box(GOLD_BORDER))
    return c


def kpi_card(ws, row, col, width, label_text, formula, fmt=FMT_CAD0, accent=TEAL, height=4):
    """A KPI card: coloured strip, small caption, big formula-driven number."""
    r1, c1, r2, c2 = row, col, row + height - 1, col + width - 1
    rect(ws, r1, c1, r2, c2, color=CARD)
    accent_bar(ws, r1, c1, c2, accent)
    cap = merge(ws, r1 + 1, c1, r1 + 1, c2)
    setc(ws, cap.coordinate, value=label_text.upper(), color=accent, size=8.5, bold=True, align=LEFT_I, fill=CARD)
    val = merge(ws, r1 + 2, c1, r2, c2)
    setc(ws, val.coordinate, value=formula, color=WHITE, size=16, bold=True, align=LEFT_I, fmt=fmt, fill=CARD)
    return val.coordinate


def col_widths(ws, widths):
    """widths: dict {column_letter_or_index: width}"""
    for k, w in widths.items():
        letter = k if isinstance(k, str) else get_column_letter(k)
        ws.column_dimensions[letter].width = w


def freeze(ws, addr):
    ws.freeze_panes = addr


def set_tab(ws, color):
    ws.sheet_properties.tabColor = color


# ════════════════════════════════════════════════════════════════════════
#  CHART STYLE HELPERS  (validated against openpyxl 3.1.5)
# ════════════════════════════════════════════════════════════════════════

def _rich(text, color=WHITE, size=1300, bold=True):
    cp = CharacterProperties(sz=size, b=bold, solidFill=color, latin=DFont(typeface=FONT_NAME))
    run = RegularTextRun(rPr=cp, t=text)
    p = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[run])
    return RichText(bodyPr=RichTextProperties(), p=[p])


def chart_title(text, color=WHITE, size=1300):
    return Title(tx=Text(rich=_rich(text, color, size, True)), overlay=False)


def axis_text(color=GREY, size=850):
    cp = CharacterProperties(sz=size, solidFill=color, latin=DFont(typeface=FONT_NAME))
    p = Paragraph(pPr=ParagraphProperties(defRPr=cp))
    return RichText(bodyPr=RichTextProperties(), p=[p])


def no_line():
    return LineProperties(noFill=True)


def thin_line(color=GRID, w=9525):
    return LineProperties(solidFill=color, w=w)


def style_chart(chart, title=None, height=8.6, width=15.5, legend="b", show_legend=True):
    """Apply the shared dark-theme look to any openpyxl chart object."""
    chart.title = chart_title(title) if title else None
    chart.style = None
    chart.graphical_properties = GraphicalProperties(solidFill=CARD, ln=no_line())
    try:
        chart.plot_area.graphicalProperties = GraphicalProperties(solidFill=CARD, ln=no_line())
    except Exception:
        pass
    if hasattr(chart, "x_axis"):
        chart.x_axis.txPr = axis_text()
        chart.x_axis.spPr = GraphicalProperties(ln=thin_line())
        chart.x_axis.delete = False
        if chart.x_axis.majorGridlines is not None:
            chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=thin_line(GRID, 6350))
    if hasattr(chart, "y_axis"):
        chart.y_axis.txPr = axis_text()
        chart.y_axis.spPr = GraphicalProperties(ln=no_line())
        chart.y_axis.delete = False
        if chart.y_axis.majorGridlines is not None:
            chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=thin_line(GRID, 6350))
    if show_legend:
        chart.legend = Legend()
        chart.legend.position = legend
        chart.legend.txPr = axis_text()
        chart.legend.overlay = False
    else:
        chart.legend = None
    chart.height = height
    chart.width = width
    return chart


def colour_series(chart, colors, line=False):
    """Assign solid fill / line colours to each series of a chart, in order."""
    for s, color in zip(chart.series, colors):
        gp = GraphicalProperties()
        if line:
            gp.line = thin_line(color, 28575)
        else:
            gp.solidFill = color
            gp.ln = LineProperties(noFill=True)
        s.graphicalProperties = gp
    return chart


def _two_stop_gradient(top_color, bottom_color):
    from openpyxl.drawing.fill import GradientFillProperties, GradientStop, LinearShadeProperties
    s1 = GradientStop(pos=0, srgbClr=top_color)
    s2 = GradientStop(pos=100000, srgbClr=bottom_color)
    return GradientFillProperties(gsLst=(s1, s2), lin=LinearShadeProperties(ang=5400000, scaled=True))


def gradient_area_series(series, top_color, bottom_color=CARD, line_color=None):
    """Give an AreaChart series a soft top-to-bottom gradient fill that fades
    into the card background (the 'gradient-style fill' look from the
    screenshots), plus a crisp solid line of the accent colour on top."""
    gp = GraphicalProperties()
    gp.gradFill = _two_stop_gradient(top_color, bottom_color)
    gp.line = thin_line(line_color or top_color, 25400)
    series.graphicalProperties = gp


def doughnut_chart(ws, anchor, data_ref, cats_ref, colors, title=None,
                   hole=62, height=8.6, width=10.5, titles_from_data=True):
    ch = DoughnutChart()
    ch.add_data(data_ref, titles_from_data=titles_from_data)
    ch.set_categories(cats_ref)
    ch.firstSliceAng = 0
    ch.holeSize = hole
    dl = DataLabelList()
    dl.showPercent = True
    dl.showCatName = False
    dl.showVal = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.numFmt = "0%"
    dl.txPr = axis_text(WHITE, 850)
    ch.dataLabels = dl
    ser = ch.series[0]
    pts = []
    for i, col in enumerate(colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties = GraphicalProperties(solidFill=col, ln=LineProperties(solidFill=NAVY, w=28575))
        pts.append(dp)
    ser.data_points = pts
    style_chart(ch, title=title, height=height, width=width, legend="b")
    ws.add_chart(ch, anchor)
    return ch


# ════════════════════════════════════════════════════════════════════════
#  SHARED CONSTANTS  (cross-sheet plan data lives here so every builder
#  function can see it without re-typing magic numbers)
# ════════════════════════════════════════════════════════════════════════

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

TRACKING_START = dt.date(2026, 6, 1)   # the workbook's 12-month tracking window starts here

# global registry of cross-sheet cell references, filled in as each sheet builds.
ctx = {}


def qs(name):
    """Quoted sheet name reference prefix, e.g.  'My Sheet'!  """
    return quote_sheetname(name) + "!"


print("helpers loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  WORKBOOK SCAFFOLD
# ════════════════════════════════════════════════════════════════════════

TAB_ORDER = [
    ("START HERE",          GREY),
    ("Dashboard",           TEAL),
    ("Income",              TEAL),
    ("Expenses",            PINK),
    ("Bills & Subscriptions", YELLOW),
    ("Debt Tracker",        PINK),
    ("Savings & Goals",     TEAL),
    ("Net Worth",           BLUE),
    ("Side Income",         PURPLE),
    ("Habit Tracker",       YELLOW),
]


def setup_sheet(ws, rows=90, cols=24, zoom=110, color=NAVY):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    paint_bg(ws, rows=rows, cols=cols, color=color)
    ws.column_dimensions["A"].width = 2.6


def build_lists_sheet(wb):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"
    data = {
        "A": ["", "✓"],                                                              # tick box options
        "B": ["Yes", "No"],                                                          # yes/no options
        "C": ["Groceries/Food", "Takeout/Dining", "Transport", "Entertainment",
              "Personal/Shopping", "Other"],                                         # discretionary categories
        "D": ["Tax refund (one-time)", "CGEB top-up (one-time)",
              "CGEB (quarterly)", "OTB (monthly)", "Other government payment"],      # gov't money types
        "E": ["Influencer campaign", "Social media management",
              "Content/UGC", "Other"],                                               # side-income categories
        "F": ["Software/Subscriptions", "Gear/Equipment", "Other"],                  # side-expense categories
        "G": ["On track", "Behind", "Ahead"],                                        # goal status options
    }
    for col, values in data.items():
        for i, v in enumerate(values, start=1):
            ws[f"{col}{i}"] = v
    ctx["lists"] = {
        "tick": f"{qs('Lists')}$A$1:$A$2",
        "yesno": f"{qs('Lists')}$B$1:$B$2",
        "disc_cat": f"{qs('Lists')}$C$1:$C$6",
        "gov_type": f"{qs('Lists')}$D$1:$D$5",
        "side_inc_cat": f"{qs('Lists')}$E$1:$E$4",
        "side_exp_cat": f"{qs('Lists')}$F$1:$F$3",
    }
    return ws


def add_dropdown(ws, cell_range, source, prompt=None, title="Pick from the list"):
    dv = DataValidation(type="list", formula1=source, allow_blank=True,
                        showDropDown=False, showInputMessage=bool(prompt),
                        showErrorMessage=True, promptTitle=title, prompt=prompt,
                        errorTitle="Please pick from the list",
                        error="Use the dropdown arrow to choose one of the listed options.")
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def define(name, sheet_title, addr):
    """Register a workbook-level defined name pointing at a single cell / range,
    and remember the raw reference too (charts need raw references)."""
    ref = f"{qs(sheet_title)}{addr}"
    ctx[name] = ref
    return ref


def add_named_ranges(wb):
    for name, ref in ctx.get("_names", {}).items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


print("scaffold helpers loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  NAMED-RANGE REGISTRY
# ════════════════════════════════════════════════════════════════════════
ctx.setdefault("_names", {})

def name_range(sheet_title, addr, name):
    ref = f"{qs(sheet_title)}{addr}"
    ctx["_names"][name] = ref
    return name

def hdr_row(ws, row, c1, headers, color=GREY, fill=CARD2, size=9.5):
    """Write a styled table header row starting at column c1."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=c1 + i)
        cell.value = h
        cell.font = fnt(color, size, True)
        cell.fill = fl(fill)
        cell.alignment = CENTER if i else LEFT_I
        cell.border = border_bottom(GRID)

def stripe(ws, r1, r2, c1, c2, color_a=CARD, color_b="262E43"):
    for idx, r in enumerate(range(r1, r2 + 1)):
        rect(ws, r, c1, r, c2, color=color_a if idx % 2 == 0 else color_b)


# ════════════════════════════════════════════════════════════════════════
#  SHEET 3 — INCOME
# ════════════════════════════════════════════════════════════════════════

def build_income(wb):
    ws = wb.create_sheet("Income")
    setup_sheet(ws, rows=78, cols=14)
    set_tab(ws, TEAL)
    col_widths(ws, {"A":2.6,"B":20,"C":15,"D":15,"E":16,"F":20,"G":3,"H":17,"I":13,"J":13,"K":13})
    title_block(ws, 2, 2, "INCOME", "Your job, your real paycheques, and government money — all in one place.", accent=TEAL)

    # ---- PAYCHECK ESTIMATOR ----
    section_header(ws, 5, 2, "Paycheque Estimator", TEAL, span=5)
    note(ws, 6, 2, "Type into the gold cells only. It estimates your 2-week take-home. The real Paycheque Log below is what the dashboard actually trusts.", span=5, height=28)

    est = [
        ("Pay rate (per hour)",        8,  21.30, FMT_CAD),
        ("Hours worked (2-week period)",9,  83,    FMT_NUM1),
        ("Overtime hours (after 44/wk — rare)", 10, 0, FMT_NUM1),
        ("Overtime rate (per hour)",   11, 31.95, FMT_CAD),
        ("Income tax % (match your stub)", 12, 0.05, FMT_PCT),
    ]
    for text, r, val, fmt in est:
        label(ws, f"B{r}", text, color=GREY)
        merge(ws, r, 2, r, 3)
        input_cell(ws, f"D{r}", val, fmt=fmt)
    name_range("Income", "$D$8", "Pay_Rate")
    name_range("Income", "$D$9", "Est_Hours")
    name_range("Income", "$D$10", "Est_OT")
    name_range("Income", "$D$11", "OT_Rate")
    name_range("Income", "$D$12", "Tax_Pct")

    calc = [
        ("Base pay",          13, "=Pay_Rate*Est_Hours", WHITE),
        ("Vacation pay (4%)",  14, "=(Pay_Rate*Est_Hours)*0.04", TEAL),
        ("Overtime pay",      15, "=Est_OT*OT_Rate", WHITE),
        ("Gross pay",         16, "=D13+D14+D15", WHITE),
        ("Less: CPP (5.95% over $134.62)", 17, "=-MAX(0,(D16-134.62)*0.0595)", PINK),
        ("Less: EI (1.63%)",  18, "=-D16*0.0163", PINK),
        ("Less: Union dues (1.57%)", 19, "=-D16*0.0157", PINK),
        ("Less: Income tax",  20, "=-D16*Tax_Pct", PINK),
    ]
    for text, r, formula, color in calc:
        label(ws, f"B{r}", text, color=GREY)
        merge(ws, r, 2, r, 3)
        value_cell(ws, f"D{r}", formula, fmt=FMT_CAD, color=color, size=10.5, align=CENTER)
    # estimated net
    rect(ws, 21, 2, 21, 4, color=CARD3)
    label(ws, "B21", "ESTIMATED TAKE-HOME (2 weeks)", color=TEAL, bold=True)
    merge(ws, 21, 2, 21, 3)
    value_cell(ws, "D21", "=D16+D17+D18+D19+D20", fmt=FMT_CAD, color=TEAL, size=13, align=CENTER)

    # sanity reference card (right)
    rect(ws, 8, 8, 16, 11, color=CARD)
    accent_bar(ws, 8, 8, 11, BLUE)
    merge(ws, 8, 8, 8, 11); setc(ws, "H8", "SANITY CHECK — MONTHLY TAKE-HOME", color=BLUE, size=9, bold=True, align=LEFT_I, fill=CARD)
    merge(ws, 9, 8, 9, 9); setc(ws,"H9","Weekly hours", color=GREY, size=9, bold=True, align=LEFT_I, fill=CARD)
    merge(ws, 9, 10, 9, 11); setc(ws,"J9","≈ Monthly", color=GREY, size=9, bold=True, align=CENTER, fill=CARD)
    ref = [("20 hrs","$1,610"),("24 hrs","$1,880"),("30 hrs","$2,265"),("35 hrs","$2,590"),("40 hrs","$2,915")]
    for i,(h,m) in enumerate(ref):
        r = 10+i
        merge(ws, r, 8, r, 9); setc(ws, f"H{r}", h, color=WHITE, size=10, align=LEFT_I, fill=CARD)
        merge(ws, r, 10, r, 11); setc(ws, f"J{r}", m, color=TEAL, size=10, bold=True, align=CENTER, fill=CARD)

    # ---- PAYCHECK LOG ----
    section_header(ws, 24, 2, "Paycheque Log — type each real paycheque here", TEAL, span=6)
    note(ws, 25, 2, "This is the most important table. Each payday, type the DATE and the NET amount from your real pay stub. The dashboard adds these up automatically.", span=6, height=28)
    hdr_row(ws, 26, 2, ["Date paid", "Hours", "Net amount (CAD)", "Notes"], fill=CARD2)
    merge(ws, 26, 5, 26, 6)
    LOG_TOP, LOG_BOT = 27, 50
    stripe(ws, LOG_TOP, LOG_BOT, 2, 6)
    preload = {27:(dt.date(2026,6,3), 24, 423.70, "First short cheque"),
               28:(dt.date(2026,6,17), 83, 1521.00, "83 hrs this period")}
    for r in range(LOG_TOP, LOG_BOT+1):
        input_cell(ws, f"B{r}", fmt=FMT_DATE, align=CENTER)
        input_cell(ws, f"C{r}", fmt=FMT_NUM1, align=CENTER)
        input_cell(ws, f"D{r}", fmt=FMT_CAD, align=CENTER)
        merge(ws, r, 5, r, 6); input_cell(ws, f"E{r}", align=LEFT_I)
        if r in preload:
            d,h,n,note_t = preload[r]
            ws[f"B{r}"].value = d; ws[f"C{r}"].value = h; ws[f"D{r}"].value = n; ws[f"E{r}"].value = note_t
    rect(ws, LOG_BOT+1, 2, LOG_BOT+1, 4, color=CARD3)
    label(ws, f"B{LOG_BOT+1}", "TOTAL LOGGED (all time)", color=TEAL, bold=True); merge(ws, LOG_BOT+1, 2, LOG_BOT+1, 3)
    value_cell(ws, f"D{LOG_BOT+1}", f"=SUM(D{LOG_TOP}:D{LOG_BOT})", fmt=FMT_CAD, color=TEAL, align=CENTER)
    name_range("Income", f"$B${LOG_TOP}:$B${LOG_BOT}", "Pay_Date")
    name_range("Income", f"$D${LOG_TOP}:$D${LOG_BOT}", "Pay_Net")

    # ---- GOVERNMENT MONEY ----
    grow = LOG_BOT + 4
    section_header(ws, grow, 2, "Government Money (extra income, added automatically)", TEAL, span=7)
    note(ws, grow+1, 2, "These top up your monthly income on the dashboard. 'Interval' is the math switch: 0 = one-time, 1 = every month, 3 = every quarter. Leave it as set.", span=7, height=28)
    hdr_row(ws, grow+2, 2, ["Payment", "Amount", "Starts", "How often", "Interval"], fill=CARD2)
    GT = grow+3; GB = grow+9
    stripe(ws, GT, GB, 2, 6)
    gov = [
        ("Tax refund (one-time)",      1211.65, dt.date(2026,6,1), "One-time",      0),
        ("CGEB top-up (one-time)",      174.50, dt.date(2026,6,1), "One-time",      0),
        ("CGEB (quarterly)",            169.75, dt.date(2026,7,1), "Every quarter", 3),
        ("OTB (monthly)",                71.95, dt.date(2026,7,1), "Every month",   1),
    ]
    for i in range(GT, GB+1):
        input_cell(ws, f"B{i}", align=LEFT_I)
        input_cell(ws, f"C{i}", fmt=FMT_CAD, align=CENTER)
        input_cell(ws, f"D{i}", fmt=FMT_MONY, align=CENTER)
        input_cell(ws, f"E{i}", align=CENTER)
        input_cell(ws, f"F{i}", fmt="0", align=CENTER)
        if i-GT < len(gov):
            nm,amt,st,freq,iv = gov[i-GT]
            ws[f"B{i}"].value=nm; ws[f"C{i}"].value=amt; ws[f"D{i}"].value=st
            ws[f"E{i}"].value=freq; ws[f"F{i}"].value=iv
    name_range("Income", f"$C${GT}:$C${GB}", "Gov_Amt")
    name_range("Income", f"$D${GT}:$D${GB}", "Gov_Start")
    name_range("Income", f"$F${GT}:$F${GB}", "Gov_Int")

    ws.print_area = "A1:N66"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("income builder loaded ok")


# 12-month tracking window (first-of-month dates), Jun 2026 .. May 2027
TRACK_MONTHS = [dt.date(2026 + (5 + i)//12, (5 + i) % 12 + 1, 1) for i in range(12)]


# ════════════════════════════════════════════════════════════════════════
#  SHEET 4 — EXPENSES
# ════════════════════════════════════════════════════════════════════════

def build_expenses(wb):
    ws = wb.create_sheet("Expenses")
    setup_sheet(ws, rows=72, cols=14)
    set_tab(ws, PINK)
    col_widths(ws, {"A":2.6,"B":22,"C":15,"D":15,"E":16,"F":22,"G":3,"H":15,"I":15,"J":15})
    title_block(ws, 2, 2, "EXPENSES", "Your fixed monthly bills, plus a takeout/spending log so you can watch it add up.", accent=PINK)

    # ---- FIXED MONTHLY BILLS ----
    section_header(ws, 5, 2, "Fixed Monthly Bills", PINK, span=5)
    note(ws, 6, 2, "Type the amount and the month it starts. 'Starts' lets a bill begin later — your India loan starts Jul 2026. Leave older bills at Jan 2020 to mean 'always on'.", span=5, height=28)
    hdr_row(ws, 7, 2, ["Bill", "Amount", "Starts", "Notes"], fill=CARD2)
    merge(ws, 7, 5, 7, 6)
    FT, FB = 8, 15
    stripe(ws, FT, FB, 2, 6)
    fixed = [
        ("Rent",            600.00, dt.date(2020,1,1), "Paid ~28th–29th of prior month"),
        ("Phone (Rogers)",  103.28, dt.date(2020,1,1), "Around the 8th–9th"),
        ("India loan payment", 334.00, dt.date(2026,7,1), "Starts Jul 2026 — see Debt sheet"),
        ("Groceries",       100.00, dt.date(2020,1,1), "Monthly budget"),
        ("Miscellaneous",    50.00, dt.date(2020,1,1), "Buffer for small stuff"),
    ]
    for i in range(FT, FB+1):
        input_cell(ws, f"B{i}", align=LEFT_I)
        input_cell(ws, f"C{i}", fmt=FMT_CAD, align=CENTER)
        input_cell(ws, f"D{i}", fmt=FMT_MONY, align=CENTER)
        merge(ws, i, 5, i, 6); input_cell(ws, f"E{i}", align=LEFT_I)
        if i-FT < len(fixed):
            nm,amt,st,nt = fixed[i-FT]
            ws[f"B{i}"].value=nm; ws[f"C{i}"].value=amt; ws[f"D{i}"].value=st; ws[f"E{i}"].value=nt
    rect(ws, FB+1, 2, FB+1, 4, color=CARD3)
    label(ws, f"B{FB+1}", "TOTAL FIXED / MONTH", color=PINK, bold=True); merge(ws, FB+1, 2, FB+1, 2)
    value_cell(ws, f"C{FB+1}", f"=SUM(C{FT}:C{FB})", fmt=FMT_CAD, color=PINK, align=CENTER)
    name_range("Expenses", f"$C${FT}:$C${FB}", "Fixed_Amt")
    name_range("Expenses", f"$D${FT}:$D${FB}", "Fixed_Start")

    # ---- DISCRETIONARY / TAKEOUT LOG ----
    section_header(ws, 18, 2, "Spending to Watch — takeout & extras log", PINK, span=6)
    note(ws, 19, 2, "Every time you spend on takeout or extras, add a line. The bar below fills up against your $100 grocery budget so you SEE it adding up.", span=6, height=28)
    hdr_row(ws, 20, 2, ["Date", "Category", "Amount", "Note"], fill=CARD2)
    merge(ws, 20, 5, 20, 6)
    DT_, DB = 21, 44
    stripe(ws, DT_, DB, 2, 6)
    for i in range(DT_, DB+1):
        input_cell(ws, f"B{i}", fmt=FMT_DATE, align=CENTER)
        input_cell(ws, f"C{i}", align=LEFT_I)
        input_cell(ws, f"D{i}", fmt=FMT_CAD, align=CENTER)
        merge(ws, i, 5, i, 6); input_cell(ws, f"E{i}", align=LEFT_I)
    ws[f"B{DT_}"].value = dt.date(2026,6,2); ws[f"C{DT_}"].value = "Takeout/Dining"
    ws[f"D{DT_}"].value = 68.90; ws[f"E{DT_}"].value = "Big takeout day"
    add_dropdown(ws, f"C{DT_}:C{DB}", ctx["lists"]["disc_cat"], "Pick a spending category")
    rect(ws, DB+1, 2, DB+1, 4, color=CARD3)
    label(ws, f"B{DB+1}", "TOTAL LOGGED (all time)", color=PINK, bold=True); merge(ws, DB+1, 2, DB+1, 3)
    value_cell(ws, f"D{DB+1}", f"=SUM(D{DT_}:D{DB})", fmt=FMT_CAD, color=PINK, align=CENTER)
    name_range("Expenses", f"$B${DT_}:$B${DB}", "Disc_Date")
    name_range("Expenses", f"$D${DT_}:$D${DB}", "Disc_Amt")

    # grocery-budget watch card
    rect(ws, 20, 8, 27, 10, color=CARD)
    accent_bar(ws, 20, 8, 10, PINK)
    merge(ws, 20, 8, 20, 10); setc(ws, "H20", "THIS MONTH vs $100 BUDGET", color=PINK, size=9, bold=True, align=LEFT_I, fill=CARD)
    merge(ws, 21, 8, 21, 10)
    setc(ws, "H21", "=SUMIFS(Disc_Amt,Disc_Date,\">=\"&Sel_Month,Disc_Date,\"<\"&EDATE(Sel_Month,1))",
         color=WHITE, size=18, bold=True, align=LEFT_I, fmt=FMT_CAD, fill=CARD)
    merge(ws, 22, 8, 22, 10); setc(ws, "H22", "spent on extras this month", color=GREY, size=9, italic=True, align=LEFT_I, fill=CARD)
    # mini data-bar against 100
    setc(ws, "H24", "Budget used", color=GREY, size=9, align=LEFT_I, fill=CARD); merge(ws,24,8,24,10)
    bar = setc(ws, "H25", "=MIN(1,H21/100)", color=WHITE, size=10, bold=True, align=LEFT_I, fmt=FMT_PCT0, fill=CARD2, border=box(GRID))
    merge(ws, 25, 8, 25, 10)
    ws.conditional_formatting.add("H25:J25",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PINK_DK, showValue=True))
    merge(ws, 26, 8, 26, 10); setc(ws, "H26", "(over 100% = past your grocery budget)", color=GREY_DK, size=8, italic=True, align=LEFT_I, fill=CARD)

    ws.print_area = "A1:N48"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("expenses builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 5 — BILLS & SUBSCRIPTIONS  (+ simple monthly calendar)
# ════════════════════════════════════════════════════════════════════════

def build_bills(wb):
    ws = wb.create_sheet("Bills & Subscriptions")
    setup_sheet(ws, rows=70, cols=22, zoom=100)
    set_tab(ws, YELLOW)
    widths = {"A":2.6,"B":22,"C":13,"D":9,"E":11,"F":12}
    for i in range(7, 19):  # G..R tick columns
        widths[get_column_letter(i)] = 6.5
    widths["S"] = 13
    col_widths(ws, widths)
    title_block(ws, 2, 2, "BILLS & SUBSCRIPTIONS", "Tick each one off when it's paid. All amounts in Canadian dollars ($).", accent=YELLOW)

    # ---- RECURRING GRID ----
    section_header(ws, 5, 2, "Recurring Bills & Subscriptions", YELLOW, span=10)
    note(ws, 6, 2, "Tick the gold box under a month once you've paid it. 'In budget?' = Yes feeds the dashboard; set it to No for bills already counted on the Expenses sheet (like rent & phone) so nothing is double-counted.", span=12, height=30)
    month_labels = [d.strftime("%b'%y") for d in TRACK_MONTHS]
    headers = ["Bill / Subscription", "Amount", "Due day", "In budget?", "Starts"] + month_labels + ["Annual"]
    hdr_row(ws, 7, 2, headers, fill=CARD2, size=8.5)
    GT, GB = 8, 19
    stripe(ws, GT, GB, 2, 19)
    preload = [
        ("Phone (Rogers)", 103.28, 9, "No",  dt.date(2026,6,1)),
        ("Claude AI",       31.64, 5, "Yes", dt.date(2026,6,1)),
    ]
    for i in range(GT, GB+1):
        input_cell(ws, f"B{i}", align=LEFT_I)
        input_cell(ws, f"C{i}", fmt=FMT_CAD, align=CENTER)
        input_cell(ws, f"D{i}", fmt="0", align=CENTER)
        input_cell(ws, f"E{i}", align=CENTER)
        input_cell(ws, f"F{i}", fmt=FMT_MONY, align=CENTER)
        for col in range(7, 19):
            tick_cell(ws, f"{get_column_letter(col)}{i}")
        value_cell(ws, f"S{i}", f"=IF(C{i}=\"\",\"\",C{i}*12)", fmt=FMT_CAD0, color=YELLOW, size=9.5, align=CENTER)
        if i-GT < len(preload):
            nm,amt,dd,bud,st = preload[i-GT]
            ws[f"B{i}"].value=nm; ws[f"C{i}"].value=amt; ws[f"D{i}"].value=dd
            ws[f"E{i}"].value=bud; ws[f"F{i}"].value=st
    add_dropdown(ws, f"E{GT}:E{GB}", ctx["lists"]["yesno"], "Should this feed the dashboard budget?")
    add_dropdown(ws, f"{get_column_letter(7)}{GT}:{get_column_letter(18)}{GB}", ctx["lists"]["tick"], "Tick when paid")
    # colour ticked boxes teal
    ws.conditional_formatting.add(f"G{GT}:R{GB}",
        FormulaRule(formula=['G8="✓"'], fill=fl(TEAL), font=fnt(NAVY, 11, True), stopIfTrue=False))
    rect(ws, GB+1, 2, GB+1, 5, color=CARD3)
    label(ws, f"B{GB+1}", "ANNUAL TOTAL (all rows)", color=YELLOW, bold=True); merge(ws, GB+1, 2, GB+1, 4)
    value_cell(ws, f"S{GB+1}", f"=SUM(S{GT}:S{GB})", fmt=FMT_CAD0, color=YELLOW, align=CENTER)
    name_range("Bills & Subscriptions", f"$C${GT}:$C${GB}", "Bill_Amt")
    name_range("Bills & Subscriptions", f"$D${GT}:$D${GB}", "Bill_DueDay")
    name_range("Bills & Subscriptions", f"$E${GT}:$E${GB}", "Bill_InBudget")
    name_range("Bills & Subscriptions", f"$F${GT}:$F${GB}", "Bill_Start")

    # ---- SHORT-TERM PLANS ----
    section_header(ws, 22, 2, "Short-term Payment Plans (finish on a date)", YELLOW, span=6)
    note(ws, 23, 2, "Interest-free plans that end. The perfume plan is pre-loaded. The 'Finish date' updates from the last dated row automatically.", span=8, height=26)
    hdr_row(ws, 24, 2, ["Plan / Item", "Date", "Amount", "Paid?"], fill=CARD2)
    PT, PB = 25, 30
    stripe(ws, PT, PB, 2, 5)
    plan = [
        ("Perfume payment plan", dt.date(2026,6,5),  60.40, "✓"),
        ("Perfume payment plan", dt.date(2026,6,19), 60.40, ""),
        ("Perfume payment plan", dt.date(2026,7,3),  60.40, ""),
        ("Perfume payment plan", dt.date(2026,7,17), 60.38, ""),
    ]
    for i in range(PT, PB+1):
        input_cell(ws, f"B{i}", align=LEFT_I)
        input_cell(ws, f"C{i}", fmt=FMT_DATE, align=CENTER)
        input_cell(ws, f"D{i}", fmt=FMT_CAD, align=CENTER)
        tick_cell(ws, f"E{i}")
        if i-PT < len(plan):
            nm,d,a,p = plan[i-PT]
            ws[f"B{i}"].value=nm; ws[f"C{i}"].value=d; ws[f"D{i}"].value=a; ws[f"E{i}"].value=p
    add_dropdown(ws, f"E{PT}:E{PB}", ctx["lists"]["tick"], "Tick when paid")
    ws.conditional_formatting.add(f"E{PT}:E{PB}",
        FormulaRule(formula=[f'E{PT}="✓"'], fill=fl(TEAL), font=fnt(NAVY, 11, True), stopIfTrue=False))
    rect(ws, PB+1, 2, PB+1, 5, color=CARD3)
    label(ws, f"B{PB+1}", "PLAN TOTAL", color=YELLOW, bold=True); merge(ws, PB+1, 2, PB+1, 2)
    value_cell(ws, f"D{PB+1}", f"=SUM(D{PT}:D{PB})", fmt=FMT_CAD, color=YELLOW, align=CENTER)
    label(ws, f"B{PB+2}", "Finishes on", color=GREY); merge(ws, PB+2, 2, PB+2, 2)
    value_cell(ws, f"D{PB+2}", f"=IF(COUNT(C{PT}:C{PB})=0,\"—\",MAX(C{PT}:C{PB}))", fmt=FMT_DATE, color=WHITE, align=CENTER)
    name_range("Bills & Subscriptions", f"$C${PT}:$C${PB}", "Plan_Date")
    name_range("Bills & Subscriptions", f"$D${PT}:$D${PB}", "Plan_Amt")

    # ---- MONTHLY CALENDAR ----
    section_header(ws, 34, 2, "Monthly Bill Calendar", YELLOW, span=8)
    label(ws, "B35", "Show month:", color=GREY); 
    input_cell(ws, "C35", dt.date(2026,6,1), fmt=FMT_MONY, align=CENTER)
    name_range("Bills & Subscriptions", "$C$35", "Cal_Month")
    title = merge(ws, 35, 5, 35, 11)
    setc(ws, "E35", '=TEXT(Cal_Month,"mmmm yyyy")', color=WHITE, size=14, bold=True, align=CENTER)
    note(ws, 36, 2, "Days with a bill due are highlighted pink (matched from the 'Due day' column above). Change the month to see another month.", span=12, height=24)
    weekdays = ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]
    for j, wd in enumerate(weekdays):
        c = 3 + j
        cell = ws.cell(row=37, column=c)
        cell.value = wd; cell.font = fnt(YELLOW, 9.5, True); cell.fill = fl(CARD2); cell.alignment = CENTER
        cell.border = border_bottom(GRID)
    CAL_TOP = 38
    for wk in range(6):
        r = CAL_TOP + wk
        ws.row_dimensions[r].height = 26
        for j in range(7):
            c = 3 + j
            off = wk*7 + j
            letter = get_column_letter(c)
            f = (f'=IF(MONTH(Cal_Month-WEEKDAY(Cal_Month)+1+{off})=MONTH(Cal_Month),'
                 f'DAY(Cal_Month-WEEKDAY(Cal_Month)+1+{off}),"")')
            cell = ws.cell(row=r, column=c)
            cell.value = f
            cell.font = fnt(GREY, 10, False); cell.fill = fl(CARD); cell.alignment = Alignment(horizontal="right", vertical="top")
            cell.border = box(GRID)
    CAL_BOT = CAL_TOP + 5
    ws.conditional_formatting.add(f"C{CAL_TOP}:I{CAL_BOT}",
        FormulaRule(formula=[f'AND(C{CAL_TOP}<>"",COUNTIF(Bill_DueDay,C{CAL_TOP})>0)'],
                    fill=fl(PINK_DK), font=fnt(WHITE, 10, True), stopIfTrue=False))

    ws.print_area = "A1:S44"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("bills builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 9 — SIDE INCOME  (freelance tracker)
# ════════════════════════════════════════════════════════════════════════

def build_side(wb):
    ws = wb.create_sheet("Side Income")
    setup_sheet(ws, rows=78, cols=18, zoom=105)
    set_tab(ws, PURPLE)
    col_widths(ws, {"A":2.6,"B":14,"C":22,"D":20,"E":14,"F":3,"G":14,"H":22,"I":20,"J":14,
                    "K":3,"L":12,"M":13,"N":13,"O":13})
    title_block(ws, 2, 2, "SIDE INCOME", "Freelance & content income — set up and ready for when you start earning.", accent=PURPLE)

    # KPI strip
    kpi_card(ws, 4, 2, 4, "Total Side Income", "=SUM(Side_Inc_Amt)", FMT_CAD0, TEAL)
    kpi_card(ws, 4, 6, 4, "Total Side Expenses", "=SUM(Side_Exp_Amt)", FMT_CAD0, PINK)
    kpi_card(ws, 4, 10, 4, "Net Profit", "=SUM(Side_Inc_Amt)-SUM(Side_Exp_Amt)", FMT_CAD0, PURPLE)

    # ---- INCOME LOG ----
    section_header(ws, 9, 2, "Income Log", TEAL, span=4)
    note(ws, 10, 2, "Log every paid gig here. Pick a category from the dropdown.", span=4, height=22)
    hdr_row(ws, 11, 2, ["Date", "Client / Source", "Category", "Amount"], fill=CARD2)
    SIT, SIB = 12, 33
    stripe(ws, SIT, SIB, 2, 5)
    for i in range(SIT, SIB+1):
        input_cell(ws, f"B{i}", fmt=FMT_DATE, align=CENTER)
        input_cell(ws, f"C{i}", align=LEFT_I)
        input_cell(ws, f"D{i}", align=LEFT_I)
        input_cell(ws, f"E{i}", fmt=FMT_CAD, align=CENTER)
    add_dropdown(ws, f"D{SIT}:D{SIB}", ctx["lists"]["side_inc_cat"], "Pick an income category")
    rect(ws, SIB+1, 2, SIB+1, 5, color=CARD3)
    label(ws, f"B{SIB+1}", "TOTAL INCOME", color=TEAL, bold=True); merge(ws, SIB+1, 2, SIB+1, 4)
    value_cell(ws, f"E{SIB+1}", f"=SUM(E{SIT}:E{SIB})", fmt=FMT_CAD, color=TEAL, align=CENTER)
    name_range("Side Income", f"$B${SIT}:$B${SIB}", "Side_Inc_Date")
    name_range("Side Income", f"$E${SIT}:$E${SIB}", "Side_Inc_Amt")

    # ---- EXPENSE LOG ----
    section_header(ws, 9, 7, "Expense Log", PINK, span=4)
    note(ws, 10, 7, "Costs tied to this work — software, gear, subscriptions.", span=4, height=22)
    hdr_row(ws, 11, 7, ["Date", "Description", "Category", "Amount"], fill=CARD2)
    for i in range(SIT, SIB+1):
        input_cell(ws, f"G{i}", fmt=FMT_DATE, align=CENTER)
        input_cell(ws, f"H{i}", align=LEFT_I)
        input_cell(ws, f"I{i}", align=LEFT_I)
        input_cell(ws, f"J{i}", fmt=FMT_CAD, align=CENTER)
    add_dropdown(ws, f"I{SIT}:I{SIB}", ctx["lists"]["side_exp_cat"], "Pick an expense category")
    rect(ws, SIB+1, 7, SIB+1, 10, color=CARD3)
    label(ws, f"G{SIB+1}", "TOTAL EXPENSES", color=PINK, bold=True); merge(ws, SIB+1, 7, SIB+1, 9)
    value_cell(ws, f"J{SIB+1}", f"=SUM(J{SIT}:J{SIB})", fmt=FMT_CAD, color=PINK, align=CENTER)
    name_range("Side Income", f"$G${SIT}:$G${SIB}", "Side_Exp_Date")
    name_range("Side Income", f"$J${SIT}:$J${SIB}", "Side_Exp_Amt")

    # ---- MONTHLY SUMMARY (drives chart) ----
    section_header(ws, 36, 2, "Monthly Summary", PURPLE, span=4)
    hdr_row(ws, 37, 12, ["Month", "Income", "Expense", "Profit"], fill=CARD2, size=9)
    MT = 38
    for k, d in enumerate(TRACK_MONTHS):
        r = MT + k
        setc(ws, f"L{r}", d, color=GREY, size=9, align=CENTER, fmt="mmm yy", fill=(CARD if k%2==0 else "262E43"))
        setc(ws, f"M{r}", f'=SUMIFS(Side_Inc_Amt,Side_Inc_Date,">="&L{r},Side_Inc_Date,"<"&EDATE(L{r},1))',
             color=TEAL, size=9, align=CENTER, fmt=FMT_CAD0, fill=(CARD if k%2==0 else "262E43"))
        setc(ws, f"N{r}", f'=SUMIFS(Side_Exp_Amt,Side_Exp_Date,">="&L{r},Side_Exp_Date,"<"&EDATE(L{r},1))',
             color=PINK, size=9, align=CENTER, fmt=FMT_CAD0, fill=(CARD if k%2==0 else "262E43"))
        setc(ws, f"O{r}", f"=M{r}-N{r}", color=WHITE, size=9, align=CENTER, fmt=FMT_CAD0, fill=(CARD if k%2==0 else "262E43"))
    MB = MT + 11

    # chart: income vs expense bars
    ch = BarChart(); ch.type = "col"
    data = Reference(ws, min_col=13, max_col=14, min_row=37, max_row=MB)
    cats = Reference(ws, min_col=12, min_row=MT, max_row=MB)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    colour_series(ch, [TEAL, PINK])
    style_chart(ch, "Side Income vs Expenses", height=8.0, width=15.0)
    ws.add_chart(ch, "B14")

    note(ws, 26, 2, "This sheet sits mostly empty until you start earning — that's fine, it's ready to go.", span=10, height=22)

    ws.print_area = "A1:O50"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("side builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  HIDDEN ENGINE — MonthData  (aggregates everything per month)
# ════════════════════════════════════════════════════════════════════════

def build_monthdata(wb):
    ws = wb.create_sheet("MonthData")
    headers = ["MonthDate","Month","Paycheque","Gov't","Side","Total Income",
               "Fixed","Extras","Bills","Total Expenses","Leftover","Cumulative",
               "Inc(shown)","Exp(shown)"]
    for j,h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h).font = fnt(WHITE, 10, True)
    for k, d in enumerate(TRACK_MONTHS):
        r = 2 + k
        A = f"A{r}"
        ws[A] = d; ws[A].number_format = FMT_MONY
        ws[f"B{r}"] = f'=TEXT(A{r},"mmm yyyy")'
        ws[f"C{r}"] = f'=SUMIFS(Pay_Net,Pay_Date,">="&A{r},Pay_Date,"<"&EDATE(A{r},1))'
        ws[f"D{r}"] = (f'=SUMPRODUCT(Gov_Amt*(Gov_Start<=A{r})*(((Gov_Int=0)*(Gov_Start=A{r}))'
                       f'+((Gov_Int>0)*(MOD((YEAR(A{r})-YEAR(Gov_Start))*12+MONTH(A{r})-MONTH(Gov_Start),'
                       f'IF(Gov_Int=0,1,Gov_Int))=0))))')
        ws[f"E{r}"] = f'=SUMIFS(Side_Inc_Amt,Side_Inc_Date,">="&A{r},Side_Inc_Date,"<"&EDATE(A{r},1))'
        ws[f"F{r}"] = f"=C{r}+D{r}+E{r}"
        ws[f"G{r}"] = f'=SUMPRODUCT(Fixed_Amt*(Fixed_Start<=A{r}))'
        ws[f"H{r}"] = f'=SUMIFS(Disc_Amt,Disc_Date,">="&A{r},Disc_Date,"<"&EDATE(A{r},1))'
        ws[f"I{r}"] = (f'=SUMPRODUCT(Bill_Amt*(Bill_InBudget="Yes")*(Bill_Start<=A{r}))'
                       f'+SUMIFS(Plan_Amt,Plan_Date,">="&A{r},Plan_Date,"<"&EDATE(A{r},1))')
        ws[f"J{r}"] = f"=G{r}+H{r}+I{r}"
        ws[f"K{r}"] = f"=F{r}-J{r}"
        # "actualized" = a month where a real paycheque has been logged.
        # Cumulative savings and the trend charts only move for those months,
        # so future un-filled months stay flat instead of diving artificially.
        if k == 0:
            ws[f"L{r}"] = f"=IF(C{r}>0,K{r},0)"
        else:
            ws[f"L{r}"] = f"=L{r-1}+IF(C{r}>0,K{r},0)"
        ws[f"M{r}"] = f'=IF(C{r}>0,F{r},NA())'
        ws[f"N{r}"] = f'=IF(C{r}>0,J{r},NA())'
        for col in range(3, 15):
            ws.cell(row=r, column=col).number_format = FMT_CAD0
    name_range("MonthData", "$B$2:$B$13", "Month_Labels")
    name_range("MonthData", "$A$2:$A$13", "Month_Dates")
    ws.sheet_state = "hidden"
    ctx["md"] = ws
    return ws


# ════════════════════════════════════════════════════════════════════════
#  SHEET 2 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════

def _md_index(col_letter):
    """INDEX into a MonthData column for the dashboard's selected month."""
    return f"=INDEX(MonthData!${col_letter}$2:${col_letter}$13,MATCH(Sel_Label,Month_Labels,0))"

def build_dashboard(wb, md):
    ws = wb.create_sheet("Dashboard")
    setup_sheet(ws, rows=78, cols=22, zoom=100)
    set_tab(ws, TEAL)
    widths = {"A":2.6}
    for i in range(2, 21):
        widths[get_column_letter(i)] = 9.2
    col_widths(ws, widths)
    title_block(ws, 2, 2, "DASHBOARD", "Your whole money picture. Pick a month below — everything updates itself.", accent=TEAL)

    # ---- MONTH PICKER ----
    rect(ws, 4, 2, 4, 8, color=CARD)
    label(ws, "B4", "VIEWING MONTH:", color=YELLOW, bold=True, size=10); merge(ws, 4, 2, 4, 3)
    input_cell(ws, "D4", "Jun 2026", align=CENTER); merge(ws, 4, 4, 4, 5)
    add_dropdown(ws, "D4", "Month_Labels", "Pick the month you want to see")
    name_range("Dashboard", "$D$4", "Sel_Label")
    # Sel_Month (real date) derived from the label — tucked away, used by other sheets
    ws["P4"] = "=INDEX(Month_Dates,MATCH(Sel_Label,Month_Labels,0))"
    ws["P4"].number_format = FMT_MONY
    name_range("Dashboard", "$P$4", "Sel_Month")
    setc(ws, "F4", "← change this to switch the whole dashboard", color=GREY_DK, size=9, italic=True, align=LEFT_I, fill=CARD)
    merge(ws, 4, 6, 4, 8)

    # register the Expenses fixed-name column for the breakdown doughnut
    name_range("Expenses", "$B$8:$B$15", "Fixed_Name")

    # ---- KPI CARDS ----
    cards = [
        ("Total Savings",        "=Sav_Total",            FMT_CAD0, TEAL),
        ("This Month Income",    _md_index("F")[1:],      FMT_CAD0, TEAL),
        ("This Month Expenses",  _md_index("J")[1:],      FMT_CAD0, PINK),
        ("Leftover This Month",  _md_index("K")[1:],      FMT_CAD0, BLUE),
        ("Emergency Fund",       "=EF_Progress",          FMT_PCT0, PURPLE),
        ("On Track for $7K?",    "=OnTrack7k",            "General",YELLOW),
    ]
    for i, (lab, formula, fmt, accent) in enumerate(cards):
        c = 2 + i*3
        kpi_card(ws, 6, c, 3, lab, "="+formula if not formula.startswith("=") else formula, fmt, accent)

    # ---- ANNUAL SUMMARY BAND ----
    section_header(ws, 11, 2, "This Year (12-month window)", GREY, span=10)
    band = [
        ("Income So Far This Year",  "=SUMPRODUCT((MonthData!C2:C13>0)*MonthData!F2:F13)", TEAL),
        ("Expenses So Far This Year","=SUMPRODUCT((MonthData!C2:C13>0)*MonthData!J2:J13)", PINK),
        ("Saved So Far This Year",   "=SUMPRODUCT((MonthData!C2:C13>0)*MonthData!K2:K13)", BLUE),
    ]
    for i,(lab,formula,accent) in enumerate(band):
        c = 2 + i*6
        kpi_card(ws, 12, c, 6, lab, formula, FMT_CAD0, accent, height=3)

    # ---- HELPER: expense breakdown buckets for selected month (lower area) ----
    HB = 60
    setc(ws, f"B{HB}", "Breakdown helper (used by the doughnut)", color=GREY_DK, size=8, italic=True)
    buckets = [
        ("Rent",       '=SUMPRODUCT((Fixed_Name="Rent")*Fixed_Amt*(Fixed_Start<=Sel_Month))'),
        ("Phone",      '=SUMPRODUCT((Fixed_Name="Phone (Rogers)")*Fixed_Amt*(Fixed_Start<=Sel_Month))'),
        ("India loan", '=SUMPRODUCT((Fixed_Name="India loan payment")*Fixed_Amt*(Fixed_Start<=Sel_Month))'),
        ("Groceries",  '=SUMPRODUCT((Fixed_Name="Groceries")*Fixed_Amt*(Fixed_Start<=Sel_Month))'),
        ("Misc",       '=SUMPRODUCT((Fixed_Name="Miscellaneous")*Fixed_Amt*(Fixed_Start<=Sel_Month))'),
        ("Subscriptions", _md_index("I")),
        ("Extras",     _md_index("H")),
    ]
    for i,(nm,formula) in enumerate(buckets):
        r = HB+1+i
        ws.cell(row=r, column=2, value=nm).font = fnt(GREY, 9)
        c = ws.cell(row=r, column=3, value=formula); c.font = fnt(GREY, 9); c.number_format = FMT_CAD0
    BK_T, BK_B = HB+1, HB+7

    # ---- CHART 1: Income vs Expenses by month ----
    section_header(ws, 16, 2, "Income vs Expenses by Month", TEAL, span=10)
    ch = BarChart(); ch.type = "col"; ch.gapWidth = 60
    s1 = Series(Reference(md, min_col=13, min_row=2, max_row=13), title="Income")
    s2 = Series(Reference(md, min_col=14, min_row=2, max_row=13), title="Expenses")
    ch.append(s1); ch.append(s2)
    ch.set_categories(Reference(md, min_col=2, min_row=2, max_row=13))
    colour_series(ch, [TEAL, PINK])
    style_chart(ch, None, height=8.6, width=17.5)
    ws.add_chart(ch, "B17")

    # ---- CHART 2: Where money goes (doughnut) ----
    section_header(ws, 16, 13, "Where My Money Goes", PINK, span=8)
    dough_cats = Reference(ws, min_col=2, min_row=BK_T, max_row=BK_B)
    dough_data = Reference(ws, min_col=3, min_row=BK_T, max_row=BK_B)
    doughnut_chart(ws, "M17", dough_data, dough_cats,
                   [PINK, BLUE, PURPLE, TEAL, YELLOW, "FF9F6B", "8AD4FF"],
                   title=None, hole=58, height=8.6, width=10.5, titles_from_data=False)

    # ---- CHART 3: Savings growing over time (area, gradient) ----
    section_header(ws, 35, 2, "Savings Growing Over Time", TEAL, span=10)
    ach = AreaChart(); ach.grouping = "standard"
    sref = Reference(md, min_col=12, min_row=1, max_row=13)
    ach.add_data(sref, titles_from_data=True)
    ach.set_categories(Reference(md, min_col=2, min_row=2, max_row=13))
    gradient_area_series(ach.series[0], TEAL, CARD, TEAL)
    style_chart(ach, None, height=8.6, width=17.5, show_legend=False)
    ws.add_chart(ach, "B36")

    # ---- GOAL PROGRESS (rings + bars) ----
    section_header(ws, 35, 13, "Goal Progress", YELLOW, span=8)
    goals = [
        ("Emergency Fund → $7,150", "=EF_Progress", PURPLE),
        ("$7,000 by Dec 2026",      "=Goal7k_Pct",  TEAL),
        ("India Trip → $5,000",     "=Trip_Pct",    BLUE),
    ]
    gr = 37
    for i,(lab, formula, accent) in enumerate(goals):
        r = gr + i*2
        rect(ws, r, 13, r+1, 20, color=CARD)
        merge(ws, r, 13, r, 18); setc(ws, (r,13), lab, color=WHITE, size=10, bold=True, align=LEFT_I, fill=CARD)
        merge(ws, r, 19, r, 20); setc(ws, (r,19), formula, color=accent, size=11, bold=True, align=CENTER, fmt=FMT_PCT0, fill=CARD)
        pcell = ws.cell(row=r+1, column=13); pcell.value = formula
        merge(ws, r+1, 13, r+1, 20)
        pcell.fill = fl(CARD2); pcell.number_format = ';;;'  # hide number, show only bar
        ws.conditional_formatting.add(f"{get_column_letter(13)}{r+1}:{get_column_letter(20)}{r+1}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=accent.replace('#',''), showValue=False))

    note(ws, 44, 13, "Each bar fills toward its goal. They move on their own as your savings and inputs change.", span=8, height=26)

    ws.print_area = "A1:T45"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("dashboard builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 7 — SAVINGS & GOALS
# ════════════════════════════════════════════════════════════════════════

def build_savings(wb):
    ws = wb.create_sheet("Savings & Goals")
    setup_sheet(ws, rows=60, cols=16, zoom=105)
    set_tab(ws, TEAL)
    col_widths(ws, {"A":2.6,"B":24,"C":15,"D":15,"E":15,"F":13,"G":16,"H":16,"I":13,"J":13})
    title_block(ws, 2, 2, "SAVINGS & GOALS", "Your accounts and the goals you're saving toward. Type into gold cells only.", accent=TEAL)

    # ---- ACCOUNTS ----
    section_header(ws, 5, 2, "Accounts", TEAL, span=6)
    note(ws, 6, 2, "Update each balance whenever you check your apps. The $25 Saven membership share is NOT real savings, so it is shown separately and left out of every total.", span=8, height=28)
    hdr_row(ws, 7, 2, ["Account", "Balance", "Rate / Auto", "This month's interest", "Notes"], fill=CARD2)
    merge(ws, 7, 6, 7, 7)
    rows = [
        ("Saven HISA (emergency fund)", 300.00, 0.0285, "rate", "=B8*C8*0+D8", "Main emergency fund"),
    ]
    # Saven
    r=8
    stripe(ws, 8, 13, 2, 7)
    input_cell(ws, "C8", 300.00, fmt=FMT_CAD); label(ws,"B8","Saven HISA (emergency fund)", WHITE, 10, True)
    input_cell(ws, "D8", 0.0285, fmt=FMT_PCT)
    value_cell(ws, "E8", "=C8*D8/12", fmt=FMT_CAD, color=TEAL, align=CENTER)
    merge(ws,8,6,8,7); label(ws,"F8","Main emergency fund — earns interest", GREY, 9)
    name_range("Savings & Goals", "$C$8", "Saven_Bal")
    name_range("Savings & Goals", "$D$8", "Saven_Rate")
    # membership share
    setc(ws,"B9","  • Membership share (not savings)", color=GREY_DK, size=9, italic=True, align=LEFT_I)
    input_cell(ws, "C9", 25.00, fmt=FMT_CAD); 
    merge(ws,9,6,9,7); label(ws,"F9","One-time, no interest, excluded from totals", GREY_DK, 9)
    setc(ws,"D9","—",color=GREY_DK,size=9,align=CENTER); setc(ws,"E9","—",color=GREY_DK,size=9,align=CENTER)
    name_range("Savings & Goals", "$C$9", "Membership")
    # WS TFSA
    label(ws,"B10","Wealthsimple TFSA (XEQT)", WHITE, 10, True)
    input_cell(ws, "C10", 100.00, fmt=FMT_CAD)
    input_cell(ws, "D10", 10.00, fmt=FMT_CAD)
    merge(ws,10,6,10,7); label(ws,"F10","Long-term (5+ yrs). $ auto-invested monthly", PURPLE, 9)
    setc(ws,"E10","auto",color=PURPLE,size=9,align=CENTER)
    name_range("Savings & Goals", "$C$10", "WS_TFSA")
    name_range("Savings & Goals", "$D$10", "WS_TFSA_Auto")
    # WS HISA
    label(ws,"B11","Wealthsimple HISA", WHITE, 10, True)
    input_cell(ws, "C11", 0.00, fmt=FMT_CAD)
    setc(ws,"D11","—",color=GREY_DK,size=9,align=CENTER); setc(ws,"E11","—",color=GREY_DK,size=9,align=CENTER)
    merge(ws,11,6,11,7); label(ws,"F11","Spare HISA", GREY, 9)
    name_range("Savings & Goals", "$C$11", "WS_HISA")
    # RBC chequing
    label(ws,"B12","RBC chequing", WHITE, 10, True)
    input_cell(ws, "C12", 500.00, fmt=FMT_CAD)
    setc(ws,"D12","—",color=GREY_DK,size=9,align=CENTER); setc(ws,"E12","—",color=GREY_DK,size=9,align=CENTER)
    merge(ws,12,6,12,7); label(ws,"F12","Everyday account (usually $300–$700)", GREY, 9)
    name_range("Savings & Goals", "$C$12", "RBC_Chq")
    # totals
    rect(ws, 13, 2, 13, 7, color=CARD3)
    label(ws,"B13","TOTAL SAVINGS (excl. membership share)", TEAL, 10, True); merge(ws,13,2,13,2)
    value_cell(ws,"C13","=Saven_Bal+WS_TFSA+WS_HISA+RBC_Chq", fmt=FMT_CAD, color=TEAL, align=CENTER)
    merge(ws,13,4,13,7); label(ws,"D13","Saven + TFSA + WS HISA + chequing", GREY, 9, align=CENTER)
    name_range("Savings & Goals", "$C$13", "Sav_Total")

    # ---- GOALS ----
    section_header(ws, 16, 2, "Goals", TEAL, span=8)
    note(ws, 17, 2, "Type your Target and your planned $/month (gold). The progress, the amount you still need each month, and the projected finish all calculate themselves.", span=9, height=28)
    hdr_row(ws, 18, 2, ["Goal","Target","Saved now","Plan $/mo","Progress","Need $/mo","Finish by","On track?"], fill=CARD2, size=9)
    stripe(ws, 19, 21, 2, 9)
    # Emergency
    label(ws,"B19","Emergency fund (3mo $3,600 → 6mo $7,150)", WHITE,9.5,True)
    input_cell(ws,"C19",7150,fmt=FMT_CAD0); value_cell(ws,"D19","=Saven_Bal",fmt=FMT_CAD0,color=TEAL,align=CENTER)
    input_cell(ws,"E19",200,fmt=FMT_CAD0)
    value_cell(ws,"F19","=D19/C19",fmt=FMT_PCT0,color=WHITE,align=CENTER)
    value_cell(ws,"G19",'="—"',fmt="General",color=GREY_DK,align=CENTER)
    value_cell(ws,"H19",'=IF(E19<=0,"add plan",EDATE(TODAY(),ROUNDUP((C19-D19)/E19,0)))',fmt=FMT_MONY,color=WHITE,align=CENTER)
    value_cell(ws,"I19",'=IF(D19>=C19,"Done",IF(D19>=3600,"6mo next","build 3mo"))',fmt="General",color=TEAL,align=CENTER)
    name_range("Savings & Goals", "$F$19", "EF_Progress")
    # $7k
    label(ws,"B20","$7,000 total savings by Dec 2026", WHITE,9.5,True)
    input_cell(ws,"C20",7000,fmt=FMT_CAD0); value_cell(ws,"D20","=Sav_Total",fmt=FMT_CAD0,color=TEAL,align=CENTER)
    input_cell(ws,"E20",300,fmt=FMT_CAD0)
    value_cell(ws,"F20","=D20/C20",fmt=FMT_PCT0,color=WHITE,align=CENTER)
    value_cell(ws,"G20",'=MAX(0,(C20-D20)/MAX(1,DATEDIF(TODAY(),DATE(2026,12,1),"m")))',fmt=FMT_CAD0,color=YELLOW,align=CENTER)
    value_cell(ws,"H20",'=DATE(2026,12,1)',fmt=FMT_MONY,color=WHITE,align=CENTER)
    value_cell(ws,"I20",'=IF(D20+E20*MAX(1,DATEDIF(TODAY(),DATE(2026,12,1),"m"))>=C20,"Yes","No")',fmt="General",color=TEAL,align=CENTER)
    name_range("Savings & Goals", "$F$20", "Goal7k_Pct")
    name_range("Savings & Goals", "$C$20", "Goal7k_Target")
    name_range("Savings & Goals", "$I$20", "OnTrack7k")
    # Trip
    label(ws,"B21","India trip fund (~Jan 2027)", WHITE,9.5,True)
    input_cell(ws,"C21",5000,fmt=FMT_CAD0); input_cell(ws,"D21",0,fmt=FMT_CAD0)
    input_cell(ws,"E21",150,fmt=FMT_CAD0)
    value_cell(ws,"F21","=D21/C21",fmt=FMT_PCT0,color=WHITE,align=CENTER)
    value_cell(ws,"G21",'=MAX(0,(C21-D21)/MAX(1,DATEDIF(TODAY(),DATE(2027,1,1),"m")))',fmt=FMT_CAD0,color=YELLOW,align=CENTER)
    value_cell(ws,"H21",'=DATE(2027,1,1)',fmt=FMT_MONY,color=WHITE,align=CENTER)
    value_cell(ws,"I21",'=IF(D21+E21*MAX(1,DATEDIF(TODAY(),DATE(2027,1,1),"m"))>=C21,"On track","Behind")',fmt="General",color=TEAL,align=CENTER)
    name_range("Savings & Goals", "$F$21", "Trip_Pct")
    name_range("Savings & Goals", "$D$21", "Trip_Current")
    # progress data bars
    for r in (19,20,21):
        ws.conditional_formatting.add(f"F{r}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=TEAL_DK, showValue=True))

    # ---- DOUGHNUT: savings split ----
    section_header(ws, 24, 2, "Where Your Savings Sit", TEAL, span=6)
    setc(ws,"B55","helper", color=GREY_DK, size=8)
    splits = [("Saven (emergency)","=Saven_Bal"),("WS TFSA (long-term)","=WS_TFSA"),
              ("WS HISA","=WS_HISA"),("RBC chequing","=RBC_Chq")]
    for i,(nm,fm) in enumerate(splits):
        ws.cell(row=56+i,column=2,value=nm).font=fnt(GREY,9)
        c=ws.cell(row=56+i,column=3,value=fm); c.font=fnt(GREY,9); c.number_format=FMT_CAD0
    doughnut_chart(ws, "B25", Reference(ws,min_col=3,min_row=56,max_row=59),
                   Reference(ws,min_col=2,min_row=56,max_row=59),
                   [TEAL, PURPLE, BLUE, YELLOW], title=None, hole=58, height=8.4, width=11, titles_from_data=False)

    ws.print_area = "A1:J45"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("savings builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 6 — DEBT TRACKER
# ════════════════════════════════════════════════════════════════════════

def build_debt(wb):
    ws = wb.create_sheet("Debt Tracker")
    setup_sheet(ws, rows=70, cols=20, zoom=100)
    set_tab(ws, PINK)
    col_widths(ws, {"A":2.6,"B":26,"C":16,"D":12,"E":13,"F":16,"G":13,"H":14,"I":3,
                    "J":14,"K":13,"L":13,"M":13,"N":13,"O":13,"P":12,"Q":14})
    title_block(ws, 2, 2, "DEBT TRACKER", "Pay the Avion in full every month. Build your emergency fund, then crush the 12% India loan.", accent=PINK)

    # ---- INDIA LOAN CONVERTER ----
    section_header(ws, 5, 2, "India Education Loan — currency converter", PINK, span=7)
    note(ws, 6, 2, "Type the rupee amounts and today's exchange rate (how many rupees per $1 CAD). The CAD figures update by themselves.", span=8, height=26)
    conv = [
        ("Still owed (₹ INR)", "C7", 1500000, "0"),
        ("Exchange rate (₹ per $1 CAD)", "C8", 62.0, "0.0"),
        ("Monthly payment (₹ INR)", "C9", 23338, "0"),
        ("Interest rate (per year)", "C10", 0.12, FMT_PCT),
    ]
    for lab, cell, val, fmt in conv:
        r=int(cell[1:]); label(ws,f"B{r}",lab,GREY); input_cell(ws,cell,val,fmt=fmt)
    name_range("Debt Tracker", "$C$7", "INR_Owed")
    name_range("Debt Tracker", "$C$8", "INR_per_CAD")
    name_range("Debt Tracker", "$C$9", "INR_Pay")
    name_range("Debt Tracker", "$C$10", "India_Rate")
    label(ws,"E7","→ in CAD:",GREY); value_cell(ws,"F7","=INR_Owed/INR_per_CAD",fmt=FMT_CAD,color=PINK,align=CENTER)
    label(ws,"E9","→ in CAD:",GREY); value_cell(ws,"F9","=INR_Pay/INR_per_CAD",fmt=FMT_CAD,color=PINK,align=CENTER)
    name_range("Debt Tracker", "$F$7", "India_CAD")
    name_range("Debt Tracker", "$F$9", "India_Pay")
    note(ws, 11, 2, "Your India loan is 12%/yr — that's expensive. Once your emergency fund is built, this is the #1 thing to attack.", span=8, height=24)

    # ---- AVION CARD ----
    section_header(ws, 5, 10, "RBC Avion Visa", BLUE, span=5)
    label(ws,"J6","Credit limit",GREY); value_cell(ws,"L6",1000,fmt=FMT_CAD0,color=WHITE,align=CENTER)
    label(ws,"J7","Interest if unpaid",GREY); value_cell(ws,"L7",0.1999,fmt=FMT_PCT,color=PINK,align=CENTER)
    label(ws,"J8","Balance carried (should be $0)",GREY); input_cell(ws,"L8",0,fmt=FMT_CAD0)
    name_range("Debt Tracker", "$L$8", "Avion_Bal")
    # warning flag
    setc(ws,"J9",'=IF(Avion_Bal>0,"⚠ PAY THIS OFF — 19.99% is costing you money!","✓ Paid in full — nice work")',
         color=WHITE, size=10, bold=True, align=LEFT_I); merge(ws,9,10,9,15)
    ws.conditional_formatting.add("J9:O9",
        FormulaRule(formula=['Avion_Bal>0'], fill=fl(PINK_DK), font=fnt(WHITE,10,True), stopIfTrue=False))
    ws.conditional_formatting.add("J9:O9",
        FormulaRule(formula=['Avion_Bal<=0'], fill=fl("1F3D34"), font=fnt(TEAL,10,True), stopIfTrue=False))

    # ---- OVERVIEW TABLE ----
    section_header(ws, 13, 2, "Debt Overview (up to 12 debts)", PINK, span=8)
    note(ws, 14, 2, "India & Avion fill in from the boxes above. For any other debt, type into the gold cells. 'Months left' and 'Debt-free date' calculate from your payment.", span=9, height=26)
    hdr_row(ws, 15, 2, ["Debt","Amount owed (CAD)","Interest","Start date","Min monthly (CAD)","Months left","Debt-free date"], fill=CARD2, size=9)
    DT_, DB = 16, 27
    stripe(ws, DT_, DB, 2, 8)
    # India row (linked)
    label(ws,f"B{DT_}","India education loan", WHITE,9.5,True)
    value_cell(ws,f"C{DT_}","=India_CAD",fmt=FMT_CAD0,color=PINK,align=CENTER)
    value_cell(ws,f"D{DT_}","=India_Rate",fmt=FMT_PCT,color=WHITE,align=CENTER)
    input_cell(ws,f"E{DT_}",dt.date(2026,7,1),fmt=FMT_MONY)
    value_cell(ws,f"F{DT_}","=India_Pay",fmt=FMT_CAD0,color=WHITE,align=CENTER)
    # Avion row (linked)
    label(ws,f"B{DT_+1}","RBC Avion Visa", WHITE,9.5,True)
    value_cell(ws,f"C{DT_+1}","=Avion_Bal",fmt=FMT_CAD0,color=PINK,align=CENTER)
    value_cell(ws,f"D{DT_+1}",0.1999,fmt=FMT_PCT,color=WHITE,align=CENTER)
    input_cell(ws,f"E{DT_+1}",dt.date(2023,12,1),fmt=FMT_MONY)
    input_cell(ws,f"F{DT_+1}",0,fmt=FMT_CAD0)
    # blank debts
    for i in range(DT_+2, DB+1):
        input_cell(ws,f"B{i}",align=LEFT_I); input_cell(ws,f"C{i}",fmt=FMT_CAD0)
        input_cell(ws,f"D{i}",fmt=FMT_PCT); input_cell(ws,f"E{i}",fmt=FMT_MONY); input_cell(ws,f"F{i}",fmt=FMT_CAD0)
    # computed months-left & payoff for all rows
    for i in range(DT_, DB+1):
        ws[f"G{i}"] = (f'=IF(N(C{i})<=0,"",IF(F{i}<=C{i}*(D{i}/12),"⚠ raise pay",'
                       f'ROUNDUP(NPER(D{i}/12,-F{i},C{i}),0)))')
        ws[f"G{i}"].number_format = "0"; ws[f"G{i}"].font = fnt(WHITE,9.5,True); ws[f"G{i}"].alignment = CENTER
        ws[f"H{i}"] = f'=IF(ISNUMBER(G{i}),EDATE(TODAY(),G{i}),"—")'
        ws[f"H{i}"].number_format = FMT_MONY; ws[f"H{i}"].font = fnt(TEAL,9.5,True); ws[f"H{i}"].alignment = CENTER
    name_range("Debt Tracker", f"$B${DT_}:$B${DB}", "Debt_Name")
    name_range("Debt Tracker", f"$C${DT_}:$C${DB}", "Debt_Amt")
    name_range("Debt Tracker", f"$G${DT_}:$G${DB}", "Debt_Months")
    # totals + debt-free date
    rect(ws, DB+1, 2, DB+1, 8, color=CARD3)
    label(ws,f"B{DB+1}","TOTAL DEBT / DEBT-FREE DATE", PINK,10,True); merge(ws,DB+1,2,DB+1,2)
    value_cell(ws,f"C{DB+1}",f"=SUM(C{DT_}:C{DB})",fmt=FMT_CAD0,color=PINK,align=CENTER)
    value_cell(ws,f"H{DB+1}",f'=IFERROR(EDATE(TODAY(),MAX(IF(ISNUMBER(G{DT_}:G{DB}),G{DT_}:G{DB}))),"—")',
               fmt=FMT_MONY,color=TEAL,align=CENTER)
    ws[f"H{DB+1}"].font = fnt(TEAL,11,True)

    # ---- DECLINING-BALANCE PROJECTION (helper for area chart) ----
    pcol = 17  # column Q
    ws.cell(row=15, column=pcol, value="Year").font = fnt(GREY_DK,8)
    ws.cell(row=15, column=pcol+1, value="Projected debt").font = fnt(GREY_DK,8)
    for y in range(0, 13):
        r = 16 + y
        ws.cell(row=r, column=pcol, value=f"Y{y}").font = fnt(GREY,9)
        n = y*12
        # India remaining after n months + other current balances held flat
        f = (f'=MAX(0,India_CAD*(1+India_Rate/12)^{n}-India_Pay*(((1+India_Rate/12)^{n})-1)/(India_Rate/12))'
             f'+SUM(C{DT_+1}:C{DB})')
        c = ws.cell(row=r, column=pcol+1, value=f); c.number_format = FMT_CAD0; c.font = fnt(PINK,9)
    PJT, PJB = 16, 28

    # ---- CHARTS ----
    # declining balance area
    section_header(ws, 30, 2, "Debt Going Down Over Time", PINK, span=7)
    ach = AreaChart()
    ach.add_data(Reference(ws, min_col=pcol+1, min_row=15, max_row=PJB), titles_from_data=True)
    ach.set_categories(Reference(ws, min_col=pcol, min_row=PJT, max_row=PJB))
    gradient_area_series(ach.series[0], PINK, CARD, PINK)
    style_chart(ach, None, height=8.4, width=14.5, show_legend=False)
    ws.add_chart(ach, "B31")

    # current-balance doughnut
    section_header(ws, 30, 10, "Current Balances", PINK, span=5)
    doughnut_chart(ws, "J31", Reference(ws,min_col=3,min_row=DT_,max_row=DT_+3),
                   Reference(ws,min_col=2,min_row=DT_,max_row=DT_+3),
                   [PINK, BLUE, PURPLE, YELLOW], title=None, hole=58, height=8.4, width=9.5, titles_from_data=False)

    # months-left bar
    section_header(ws, 30, 15, "Months Left", BLUE, span=5)
    bch = BarChart(); bch.type = "col"
    bch.add_data(Reference(ws, min_col=7, min_row=DT_, max_row=DT_+3), titles_from_data=False)
    bch.set_categories(Reference(ws, min_col=2, min_row=DT_, max_row=DT_+3))
    colour_series(bch, [BLUE])
    style_chart(bch, None, height=8.4, width=9.5, show_legend=False)
    ws.add_chart(bch, "O31")

    ws.print_area = "A1:P46"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("debt builder loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 8 — NET WORTH
# ════════════════════════════════════════════════════════════════════════

def build_networth(wb, md):
    ws = wb.create_sheet("Net Worth")
    setup_sheet(ws, rows=66, cols=20, zoom=100)
    set_tab(ws, BLUE)
    widths = {"A":2.6}
    for i in range(2,21): widths[get_column_letter(i)] = 9.2
    col_widths(ws, widths)
    title_block(ws, 2, 2, "NET WORTH TRACKER", "Everything you own minus everything you owe. It updates from your other sheets.", accent=BLUE)

    # KPI cards
    kpi_card(ws, 4, 2, 5, "Total Net Worth", "=Sav_Total-(India_CAD+Avion_Bal)", FMT_CAD0, TEAL)
    kpi_card(ws, 4, 7, 5, "Net Worth Goal", "=NW_Goal", FMT_CAD0, YELLOW)
    kpi_card(ws, 4, 12, 5, "Amount to Goal", "=NW_Goal-(Sav_Total-(India_CAD+Avion_Bal))", FMT_CAD0, BLUE)
    label(ws,"B9","Set your custom net worth goal →", GREY); input_cell(ws,"E9",25000,fmt=FMT_CAD0)
    name_range("Net Worth", "$E$9", "NW_Goal")

    # ---- HELPER: assets / liabilities / net worth by month ----
    HT = 52
    ws.cell(row=HT-1, column=2, value="helper (assets vs liabilities by month)").font = fnt(GREY_DK,8)
    hdr = ["Month","Assets","Liabilities","Net worth"]
    for j,h in enumerate(hdr): ws.cell(row=HT-1, column=12+j, value=h).font = fnt(GREY_DK,8)
    for k in range(12):
        r = HT + k; idx = k+1
        ws.cell(row=r, column=12, value=f'=MonthData!B{2+k}').font = fnt(GREY,8)
        a = ws.cell(row=r, column=13, value=f'=Sav_Total+MonthData!L{2+k}'); a.number_format=FMT_CAD0; a.font=fnt(TEAL,8)
        l = ws.cell(row=r, column=14, value=f'=MAX(0,(India_CAD+Avion_Bal)-India_Pay*{idx})'); l.number_format=FMT_CAD0; l.font=fnt(PINK,8)
        n = ws.cell(row=r, column=15, value=f'=M{r}-N{r}'); n.number_format=FMT_CAD0; n.font=fnt(WHITE,8)
    HB = HT + 11

    # ---- assets & liabilities split helpers ----
    AT = 60
    asset_rows = [("Saven","=Saven_Bal"),("WS TFSA","=WS_TFSA"),("RBC chequing","=RBC_Chq"),("WS HISA","=WS_HISA")]
    for i,(nm,fm) in enumerate(asset_rows):
        ws.cell(row=AT+i, column=2, value=nm).font=fnt(GREY,8)
        c=ws.cell(row=AT+i, column=3, value=fm); c.number_format=FMT_CAD0; c.font=fnt(GREY,8)
    liab_rows = [("India loan","=India_CAD"),("Avion card","=Avion_Bal")]
    for i,(nm,fm) in enumerate(liab_rows):
        ws.cell(row=AT+i, column=5, value=nm).font=fnt(GREY,8)
        c=ws.cell(row=AT+i, column=6, value=fm); c.number_format=FMT_CAD0; c.font=fnt(GREY,8)

    # ---- CHARTS ----
    section_header(ws, 12, 2, "Net Worth Over Time", TEAL, span=8)
    bch = BarChart(); bch.type="col"
    bch.add_data(Reference(ws, min_col=15, min_row=HT-1, max_row=HB), titles_from_data=True)
    bch.set_categories(Reference(ws, min_col=12, min_row=HT, max_row=HB))
    colour_series(bch, [TEAL]); style_chart(bch, None, height=8.6, width=15.5, show_legend=False)
    ws.add_chart(bch, "B13")

    section_header(ws, 12, 11, "Assets Distribution", BLUE, span=5)
    doughnut_chart(ws, "K13", Reference(ws,min_col=3,min_row=AT,max_row=AT+3),
                   Reference(ws,min_col=2,min_row=AT,max_row=AT+3),
                   [TEAL, PURPLE, BLUE, YELLOW], title=None, hole=58, height=8.6, width=9.5, titles_from_data=False)

    section_header(ws, 31, 2, "Assets vs Liabilities", TEAL, span=8)
    ach = AreaChart()
    a1 = Series(Reference(ws, min_col=13, min_row=HT, max_row=HB), title="Assets")
    a2 = Series(Reference(ws, min_col=14, min_row=HT, max_row=HB), title="Liabilities")
    ach.append(a1); ach.append(a2)
    ach.set_categories(Reference(ws, min_col=12, min_row=HT, max_row=HB))
    gradient_area_series(ach.series[0], TEAL, CARD, TEAL)
    gradient_area_series(ach.series[1], PINK, CARD, PINK)
    style_chart(ach, None, height=8.6, width=15.5)
    ws.add_chart(ach, "B32")

    section_header(ws, 31, 11, "Liabilities Distribution", PINK, span=5)
    doughnut_chart(ws, "K32", Reference(ws,min_col=6,min_row=AT,max_row=AT+1),
                   Reference(ws,min_col=5,min_row=AT,max_row=AT+1),
                   [PINK, YELLOW], title=None, hole=58, height=8.6, width=9.5, titles_from_data=False)

    ws.print_area = "A1:T48"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# ════════════════════════════════════════════════════════════════════════
#  SHEET 10 — HABIT TRACKER
# ════════════════════════════════════════════════════════════════════════

def build_habit(wb):
    ws = wb.create_sheet("Habit Tracker")
    setup_sheet(ws, rows=55, cols=40, zoom=90)
    set_tab(ws, YELLOW)
    widths = {"A":2.6, "B":20}
    for i in range(3, 34): widths[get_column_letter(i)] = 3.4   # 31 days
    widths["AH"] = 11
    col_widths(ws, widths)
    title_block(ws, 2, 2, "HABIT TRACKER", "Tick a box each day you do the habit. The % and charts update on their own.", accent=YELLOW)

    label(ws,"B4","Show month:", GREY); input_cell(ws,"C4",dt.date(2026,6,1),fmt=FMT_MONY,align=CENTER); merge(ws,4,3,4,5)
    name_range("Habit Tracker", "$C$4", "Habit_Month")
    setc(ws,"G4",'=TEXT(Habit_Month,"mmmm yyyy")', color=WHITE, size=13, bold=True, align=LEFT_I); merge(ws,4,7,4,12)
    note(ws, 5, 2, "Days past the end of the month don't count. The % uses the real number of days in the month you picked.", span=14, height=22)

    # day header
    setc(ws,"B7","Habit",color=GREY,size=9.5,bold=True,align=LEFT_I)
    ws.cell(row=7, column=2).fill = fl(CARD2); ws.cell(row=7,column=2).border = border_bottom(GRID)
    for d in range(1,32):
        c = 2+d
        cell = ws.cell(row=7, column=c, value=d); cell.font=fnt(GREY,8,True); cell.fill=fl(CARD2)
        cell.alignment = CENTER; cell.border = border_bottom(GRID)
    pc = ws.cell(row=7, column=34, value="% done"); pc.font=fnt(YELLOW,9,True); pc.fill=fl(CARD2); pc.alignment=CENTER

    HT_, HB = 8, 15
    habits = ["Gym / workout", "Post content", "No-spend day"]
    stripe(ws, HT_, HB, 2, 34)
    for i in range(HT_, HB+1):
        input_cell(ws, f"B{i}", align=LEFT_I)
        if i-HT_ < len(habits): ws[f"B{i}"].value = habits[i-HT_]
        for d in range(1,32):
            c = 2+d
            tick_cell(ws, f"{get_column_letter(c)}{i}")
        # % done = ticks / days in month
        pct = ws.cell(row=i, column=34,
            value=f'=IF(B{i}="","",COUNTIF(C{i}:AG{i},"✓")/DAY(EOMONTH(Habit_Month,0)))')
        pct.number_format = FMT_PCT0; pct.font = fnt(TEAL,9.5,True); pct.alignment = CENTER
    add_dropdown(ws, f"C{HT_}:AG{HB}", ctx["lists"]["tick"], "Tick the day you did it")
    ws.conditional_formatting.add(f"C{HT_}:AG{HB}",
        FormulaRule(formula=['C8="✓"'], fill=fl(TEAL), font=fnt(NAVY,10,True), stopIfTrue=False))
    for r in range(HT_, HB+1):
        ws.conditional_formatting.add(f"AH{r}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=TEAL_DK, showValue=True))

    # daily totals helper (for area chart)
    DR = 40
    ws.cell(row=DR-1, column=2, value="daily totals helper").font = fnt(GREY_DK,8)
    for d in range(1,32):
        c = 2+d
        ws.cell(row=DR, column=c, value=d).font = fnt(GREY_DK,8)
        t = ws.cell(row=DR+1, column=c, value=f'=COUNTIF({get_column_letter(c)}{HT_}:{get_column_letter(c)}{HB},"✓")')
        t.font = fnt(GREY,8)

    # CHART: % done per habit (bar)
    section_header(ws, 17, 2, "Completion % by Habit", TEAL, span=8)
    bch = BarChart(); bch.type="col"
    bch.add_data(Reference(ws, min_col=34, min_row=HT_, max_row=HB), titles_from_data=False)
    bch.set_categories(Reference(ws, min_col=2, min_row=HT_, max_row=HB))
    colour_series(bch, [TEAL]); style_chart(bch, None, height=8.0, width=15.0, show_legend=False)
    bch.y_axis.numFmt = "0%"; bch.y_axis.majorGridlines = bch.y_axis.majorGridlines
    ws.add_chart(bch, "B18")

    # CHART: daily activity area
    section_header(ws, 17, 12, "Daily Activity", PURPLE, span=8)
    ach = AreaChart()
    ach.add_data(Reference(ws, min_col=3, max_col=33, min_row=DR+1, max_row=DR+1), from_rows=True, titles_from_data=False)
    ach.set_categories(Reference(ws, min_col=3, max_col=33, min_row=DR, max_row=DR))
    gradient_area_series(ach.series[0], PURPLE, CARD, PURPLE)
    style_chart(ach, None, height=8.0, width=15.0, show_legend=False)
    ws.add_chart(ach, "L18")

    ws.print_area = "A1:AH34"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


print("networth + habit builders loaded ok")


# ════════════════════════════════════════════════════════════════════════
#  SHEET 1 — START HERE
# ════════════════════════════════════════════════════════════════════════

def build_starthere(wb):
    ws = wb.create_sheet("START HERE")
    setup_sheet(ws, rows=70, cols=16, zoom=110)
    set_tab(ws, GREY)
    col_widths(ws, {"A":2.6,"B":26,"C":15,"D":15,"E":15,"F":15,"G":15,"H":15})
    title_block(ws, 2, 2, "START HERE", "Joel's money workbook — built so you never have to touch a formula.", accent=TEAL)

    # the one rule + sample
    rect(ws, 5, 2, 8, 8, color=CARD)
    accent_bar(ws, 5, 2, 8, TEAL)
    setc(ws,"B6","THE ONE RULE", color=TEAL, size=12, bold=True, align=LEFT_I, fill=CARD); merge(ws,6,2,6,8)
    setc(ws,"B7","You only ever type into the gold cells. Everything else fills itself in. This is a gold cell →",
         color=GREY, size=10.5, align=LEFT_I, fill=CARD); merge(ws,7,2,7,6)
    input_cell(ws,"H7","type here"); 
    setc(ws,"B8","If a cell isn't gold, leave it alone — it's doing math for you.", color=GREY_DK, size=9.5, italic=True, align=LEFT_I, fill=CARD); merge(ws,8,2,8,8)

    # per-sheet guide
    section_header(ws, 10, 2, "What each tab is for", TEAL, span=7)
    guide = [
        ("Dashboard", "Your overview. Pick a month from the gold dropdown at the top — all the cards and charts update.", TEAL),
        ("Income", "Log each real paycheque (date + net $). Government money (tax refund, CGEB, OTB) is set up to add itself.", TEAL),
        ("Expenses", "Your fixed bills, plus a takeout log so you can watch extra spending against your $100 grocery budget.", PINK),
        ("Bills & Subscriptions", "Tick a box when each bill is paid. The calendar shows which day each bill lands on.", YELLOW),
        ("Debt Tracker", "India loan converts from rupees automatically. Pay the Avion card in full — it warns you if you don't.", PINK),
        ("Savings & Goals", "Update your account balances. Set goal targets + monthly plans; progress bars do the rest.", TEAL),
        ("Net Worth", "Everything you own minus everything you owe. Pulls straight from your other tabs.", BLUE),
        ("Side Income", "Ready for when you start freelancing — log income & costs, see your profit.", PURPLE),
        ("Habit Tracker", "Tick a box each day. The % and charts show how you're doing.", YELLOW),
    ]
    r = 11
    for nm, desc, col in guide:
        rect(ws, r, 2, r, 8, color=CARD)
        setc(ws, (r,2), nm, color=col, size=10.5, bold=True, align=LEFT_I, fill=CARD); merge(ws, r, 2, r, 3)
        setc(ws, (r,4), desc, color=GREY, size=9.5, align=LEFT_I, fill=CARD); merge(ws, r, 4, r, 8)
        ws.row_dimensions[r].height = 20
        r += 1

    # update cadence
    section_header(ws, r+1, 2, "Your simple routine — which gold cells to update, and when", YELLOW, span=8)
    rr = r+2
    blocks = [
        ("EVERY WEEK", TEAL, [
            "Income tab → add any paycheque you got (date + net $).",
            "Expenses tab → add takeout/extra spends as they happen.",
            "Habit tab → tick your daily boxes.",
        ]),
        ("EVERY MONTH", BLUE, [
            "Savings tab → update each account balance from your apps.",
            "Bills tab → tick off bills as you pay them.",
            "Debt tab → update the Avion balance (aim for $0).",
            "Dashboard → switch the month dropdown to the new month.",
        ]),
        ("ONE-TIME SETUP (already done for you)", PURPLE, [
            "Pay rate, deduction %, fixed bill amounts.",
            "Goal targets ($7,150 emergency, $7,000 by Dec, $5,000 trip).",
            "India loan rupee amount + exchange rate, interest rates.",
            "Just tweak these if your real numbers change.",
        ]),
    ]
    for header, col, items in blocks:
        rect(ws, rr, 2, rr+len(items), 8, color=CARD)
        accent_bar(ws, rr, 2, 8, col)
        setc(ws,(rr,2), header, color=col, size=11, bold=True, align=LEFT_I, fill=CARD); merge(ws, rr, 2, rr, 8)
        for i, it in enumerate(items):
            setc(ws,(rr+1+i,2), "•  "+it, color=GREY, size=9.5, align=LEFT_I, fill=CARD); merge(ws, rr+1+i, 2, rr+1+i, 8)
        rr += len(items) + 2

    ws.print_area = f"A1:H{rr}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# ════════════════════════════════════════════════════════════════════════
#  MAIN ASSEMBLY
# ════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    build_lists_sheet(wb)
    build_income(wb)
    build_expenses(wb)
    build_bills(wb)
    build_side(wb)
    build_savings(wb)
    build_debt(wb)
    md = build_monthdata(wb)
    build_dashboard(wb, md)
    build_networth(wb, md)
    build_habit(wb)
    build_starthere(wb)

    # register every named range
    add_named_ranges(wb)

    # order the tabs the way Joel reads them
    order = ["START HERE","Dashboard","Income","Expenses","Bills & Subscriptions",
             "Debt Tracker","Savings & Goals","Net Worth","Side Income","Habit Tracker",
             "Lists","MonthData"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.active = order.index("Dashboard")

    # force Excel/Sheets to recalculate everything on open
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)

    out = "Personal_Finance_Workbook.xlsx"
    wb.save(out)
    print("SAVED:", out)
    return out


if __name__ == "__main__":
    main()
